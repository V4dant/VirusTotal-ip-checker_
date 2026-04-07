"""
VirusTotal IP Checker  (with API limit protection)
===================================================
HOW TO USE EVERY DAY:
  1. Paste your IPs into IP_LIST below
  2. Open Terminal and run:  python3 vt_ip_checker.py
  3. Excel is saved automatically

IF LIMIT RUNS OUT MID-RUN:
  - Script saves progress to  progress.json  automatically
  - Next day just run the script again — it resumes from where it stopped
  - When all IPs are done, progress.json is deleted automatically

PRESS Ctrl+C AT ANY TIME to stop and save Excel with IPs done so far.

FIRST TIME ONLY — install libraries:
  pip install requests openpyxl
"""

import time
import json
import os
import requests
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

#  1. PASTE YOUR API KEYS  (add as many as you have)

VT_API_KEYS = [
    "your_first_api_key_here",
    "your_second_api_key_here"
]
current_key_index = 0

#  2. PASTE YOUR IPs HERE  (one per line)

IP_LIST = """
ip_address_1
ip_address_2
""".strip()

#  SETTINGS

REQUESTS_PER_MINUTE = 120
PROGRESS_FILE       = "progress.json"
from datetime import datetime
OUTPUT_FILE = f"IP_Threat_Analysis_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

# Global variables so Ctrl+C can access them
ips         = []
results_map = {}


#  PROGRESS  (save & resume)

def load_progress(ips):
    if not os.path.exists(PROGRESS_FILE):
        return {}
    try:
        with open(PROGRESS_FILE, "r") as f:
            saved = json.load(f)
        if saved.get("ip_list") != ips:
            print("  Detected a NEW IP list — starting fresh.\n")
            return {}
        results = saved.get("results", {})
        print(f"  Resuming from previous run — {len(results)} IPs already done.\n")
        return results
    except Exception:
        return {}


def save_progress(ips, results):
    with open(PROGRESS_FILE, "w") as f:
        json.dump({"ip_list": ips, "results": results}, f)


def clear_progress():
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)


#  VIRUSTOTAL LOOKUP

def check_ip(ip):
    global current_key_index
    url     = f"https://www.virustotal.com/api/v3/ip_addresses/{ip}"
    headers = {"x-apikey": VT_API_KEYS[current_key_index]}

    try:
        response = requests.get(url, headers=headers, timeout=15)

        if response.status_code == 200:
            stats     = response.json()["data"]["attributes"]["last_analysis_stats"]
            malicious = stats.get("malicious", 0)
            total     = sum(stats.values())
            return malicious, total

        elif response.status_code == 404:
            return None, None

        elif response.status_code == 204:
            if current_key_index + 1 < len(VT_API_KEYS):
                current_key_index += 1
                print(f"  Quota hit — switching to API key {current_key_index + 1}")
                return check_ip(ip)
            else:
                print("  All API keys exhausted for today.")
                return "LIMIT", None

        elif response.status_code == 429:
            print("  Per-minute rate limit — waiting 60s...")
            time.sleep(60)
            headers  = {"x-apikey": VT_API_KEYS[current_key_index]}
            response = requests.get(url, headers=headers, timeout=15)
            if response.status_code == 200:
                stats     = response.json()["data"]["attributes"]["last_analysis_stats"]
                malicious = stats.get("malicious", 0)
                total     = sum(stats.values())
                return malicious, total
            else:
                print(f"  Still failing after wait — skipping")
                return "ERROR", None

        else:
            print(f"  HTTP {response.status_code}")
            return "ERROR", None

    except requests.exceptions.Timeout:
        print("  Timeout — skipping")
        return "ERROR", None

    except Exception as e:
        print(f"  Error: {e}")
        return "ERROR", None


#  EXCEL BUILDER

def build_excel(ips, results_map):
    wb = Workbook()
    ws = wb.active
    ws.title = "IP Threat Analysis"

    hdr_font   = Font(name="Calibri", bold=False, color="FFFFFF", size=12)
    hdr_fill   = PatternFill("solid", fgColor="0899A8")
    plain_font = Font(name="Calibri", color="000000", size=12)
    center     = Alignment(horizontal="center")

    for col, title in enumerate(["Source Address", "Status", "VT Rating"], 1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = hdr_font
        c.fill = hdr_fill

    for row_num, ip in enumerate(ips, start=2):
        result = results_map.get(ip, "PENDING")

        if result == "PENDING" or result == "LIMIT":
            status = "Pending (limit hit)"
            rating = ""
        elif result == "ERROR" or result == (None, None):
            status = "Unknown"
            rating = ""
        else:
            malicious, total = result
            if malicious == 0:
                status = "Clean"
                rating = f"0//{total}"
            else:
                status = "Malicious"
                rating = f"{malicious}//{total}"

        a = ws.cell(row=row_num, column=1, value=ip)
        a.font = plain_font

        b = ws.cell(row=row_num, column=2, value=status)
        b.font = plain_font; b.alignment = center

        c = ws.cell(row=row_num, column=3, value=rating)
        c.font = plain_font; c.alignment = center

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = "A2"
    
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = border

    ws2 = wb.create_sheet("Summary")
    mal     = sum(1 for ip in ips if isinstance(results_map.get(ip), tuple) and results_map[ip][0] > 0)
    cln     = sum(1 for ip in ips if isinstance(results_map.get(ip), tuple) and results_map[ip][0] == 0)
    unk     = sum(1 for ip in ips if results_map.get(ip) in ["ERROR", (None, None)])
    pending = sum(1 for ip in ips if results_map.get(ip) in ["PENDING", "LIMIT", None])

    ws2["A1"] = "Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14)

    summary_rows = [
        ("Total IPs",  len(ips), "000000"),
        ("Malicious",  mal,      "C0392B"),
        ("Clean",      cln,      "1E8449"),
        ("Unknown",    unk,      "888888"),
        ("Pending",    pending,  "B7860B"),
    ]
    for i, (label, val, color) in enumerate(summary_rows, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True, color=color, size=11)
        ws2.cell(row=i, column=2, value=val).font   = Font(name="Arial", color=color, size=11)

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 10

    wb.save(OUTPUT_FILE)


def print_summary():
    mal     = sum(1 for ip in ips if isinstance(results_map.get(ip), tuple) and results_map[ip][0] is not None and results_map[ip][0] > 0)
    cln     = sum(1 for ip in ips if isinstance(results_map.get(ip), tuple) and results_map[ip][0] == 0)
    pending = sum(1 for ip in ips if results_map.get(ip) in ["LIMIT", None])
    print(f"  Malicious : {mal}")
    print(f"  Clean     : {cln}")
    print(f"  Pending   : {pending}")
    print(f"  File      : {OUTPUT_FILE}\n")

#  MAIN

def main():
    global ips, results_map

    if not VT_API_KEYS or VT_API_KEYS[0] == "your_first_api_key_here":
        print("\nERROR: Set your API keys in VT_API_KEYS first.\n")
        return

    ips = [line.strip() for line in IP_LIST.splitlines() if line.strip()]
    if not ips:
        print("\nERROR: No IPs found in IP_LIST.\n")
        return

    delay = 60 / REQUESTS_PER_MINUTE

    print(f"\nVirusTotal IP Checker")
    print(f"{'─'*44}")
    print(f"  Total IPs    : {len(ips)}")
    print(f"  API keys     : {len(VT_API_KEYS)}")
    print(f"  Rate limit   : {REQUESTS_PER_MINUTE} req/min")
    print(f"  Est. time    : ~{round(len(ips) * delay / 60, 1)} minutes")
    print(f"  Tip          : Press Ctrl+C anytime to stop and save Excel")
    print(f"{'─'*44}\n")

    results_map = load_progress(ips)

    limit_hit = False

    for idx, ip in enumerate(ips, start=1):

        if ip in results_map:
            print(f"[{idx:>4}/{len(ips)}]  {ip:<22}  (already done — skipping)")
            continue

        if limit_hit:
            results_map[ip] = "LIMIT"
            continue

        print(f"[{idx:>4}/{len(ips)}]  {ip:<22}", end="  ", flush=True)

        result = check_ip(ip)

        if result == ("LIMIT", None):
            print("DAILY LIMIT REACHED — saving progress and stopping.")
            results_map[ip] = "LIMIT"
            limit_hit = True
            save_progress(ips, results_map)
            continue

        elif result == ("ERROR", None):
            print("error — marked as Unknown")
            results_map[ip] = "ERROR"

        elif result == (None, None):
            print("not found in VT")
            results_map[ip] = (None, None)

        else:
            malicious, total = result
            results_map[ip]  = (malicious, total)
            if malicious == 0:
                print("Clean")
            else:
                print(f"MALICIOUS  {malicious}//{total}")

        save_progress(ips, results_map)

        if idx < len(ips) and not limit_hit:
            time.sleep(delay)

    print(f"\n{'─'*44}")

    if limit_hit:
        remaining = sum(1 for ip in ips if results_map.get(ip) in ["LIMIT", None])
        print(f"  Limit hit! {remaining} IPs still pending.")
        print(f"  Run the script again tomorrow to finish.\n")
    else:
        print(f"  All IPs checked!")
        clear_progress()

    print(f"  Building Excel → {OUTPUT_FILE}")
    build_excel(ips, results_map)
    print(f"\n  Done!")
    print_summary()


#  ENTRY POINT

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print(f"\n\n  Stopped by user — building Excel with completed IPs...")
        if ips and results_map is not None:
            build_excel(ips, results_map)
            save_progress(ips, results_map)
            print_summary()
        else:
            print("  Nothing to save.\n")